import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from io import StringIO
import openpyxl
import sys
import traceback
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np

class SalesUpdateTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Sales Update and Analysis Tool")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 600)
        
        self.product_mapping = {
            'PCI': 'Pepsi',
            'GAT': 'Gatorade',
            'LIP': 'Lipton',
            'JUC': 'Juice',
            'WTR': 'Water'
        }
        
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.main_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.main_tab, text="Sales Update")
        
        self.analysis_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.analysis_tab, text="Analysis")
        
        self.create_main_widgets()
        self.create_analysis_widgets()
        
        self.excel_path = None
        self.analysis_data = {}
        
        self.root.bind("<Configure>", self.on_window_configure)
        
    def on_window_configure(self, event):
        if event.widget == self.root:
            if hasattr(self, 'analysis_display'):
                self.analysis_display.configure(width=event.width-20)
    
    def create_main_widgets(self):
        main_frame = ttk.Frame(self.main_tab, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        file_frame = ttk.LabelFrame(main_frame, text="Excel File Selection", padding="10")
        file_frame.pack(fill="x", padx=5, pady=5)
        
        self.file_label = ttk.Label(file_frame, text="No file selected", wraplength=500)
        self.file_label.pack(side="left", padx=5, fill="x", expand=True)
        
        select_button = ttk.Button(file_frame, text="Select Excel File", command=self.select_excel_file)
        select_button.pack(side="right", padx=5)
        
        input_frame = ttk.LabelFrame(main_frame, text="Sales Data Input", padding="10")
        input_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        ttk.Label(input_frame, text="Paste sales data:").pack(anchor="w")
        
        text_frame = ttk.Frame(input_frame)
        text_frame.pack(fill="both", expand=True, pady=5)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side="right", fill="y")
        
        self.paste_text = tk.Text(text_frame, height=10, yscrollcommand=scrollbar.set)
        self.paste_text.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.paste_text.yview)
        
        product_frame = ttk.Frame(input_frame)
        product_frame.pack(fill="x", pady=5)
        
        ttk.Label(product_frame, text="Select Product:").pack(side="left", padx=5)
        
        self.product_var = tk.StringVar(value="All")
        self.product_combo = ttk.Combobox(product_frame, textvariable=self.product_var)
        self.product_combo['values'] = ['All'] + list(self.product_mapping.keys())
        self.product_combo.pack(side="left", padx=5)
        
        update_button = ttk.Button(input_frame, text="Update Excel", command=self.update_excel)
        update_button.pack(pady=10)
        
        status_frame = ttk.LabelFrame(main_frame, text="Status", padding="10")
        status_frame.pack(fill="x", padx=5, pady=5)
        
        status_scroll = ttk.Scrollbar(status_frame)
        status_scroll.pack(side="right", fill="y")
        
        self.status_text = tk.Text(status_frame, height=5, wrap=tk.WORD, yscrollcommand=status_scroll.set)
        self.status_text.pack(fill="x", expand=True)
        status_scroll.config(command=self.status_text.yview)
        
    def create_analysis_widgets(self):
        control_frame = ttk.Frame(self.analysis_tab, padding="5")
        control_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(control_frame, text="Select Product:").pack(side="left", padx=5)
        self.analysis_product_var = tk.StringVar(value="All")
        self.analysis_product_combo = ttk.Combobox(control_frame, 
                                                 textvariable=self.analysis_product_var,
                                                 values=['All'] + list(self.product_mapping.keys()))
        self.analysis_product_combo.pack(side="left", padx=5)
        
        ttk.Button(control_frame, text="Analyze", command=self.perform_analysis).pack(side="left", padx=5)
        
        self.analysis_display = ttk.Frame(self.analysis_tab)
        self.analysis_display.pack(fill="both", expand=True, padx=5, pady=5)
    
    def select_excel_file(self):
        try:
            self.excel_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if self.excel_path:
                self.file_label.config(text=f"Selected: {self.excel_path}")
                self.log_status("Excel file selected successfully")
        except Exception as e:
            self.handle_error("Error selecting file", e)
            
    def log_status(self, message):
        try:
            self.status_text.insert(tk.END, f"{message}\n")
            self.status_text.see(tk.END)
            self.root.update_idletasks()
        except Exception as e:
            print(f"Error logging status: {str(e)}")
            
    def handle_error(self, message, error):
        error_msg = f"{message}: {str(error)}"
        messagebox.showerror("Error", error_msg)
        self.log_status(f"Error: {error_msg}")
    
    def parse_sales_data(self):
        try:
            pasted_data = self.paste_text.get("1.0", tk.END).strip()
            if not pasted_data:
                raise ValueError("No data pasted")
            
            lines = pasted_data.split('\n')
            
            header_idx = -1
            for i, line in enumerate(lines):
                if 'GAT' in line and 'PCI' in line and 'WTR' in line:
                    header_idx = i
                    break
            
            if header_idx == -1:
                raise ValueError("Could not find product codes in the data")
            
            cleaned_data = '\n'.join(lines[header_idx:])
            
            df = pd.read_csv(StringIO(cleaned_data), sep='\t')
            
            df.set_index(df.columns[0], inplace=True)
            
            for col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
            
            df.fillna(0, inplace=True)
            
            return df
            
        except Exception as e:
            raise Exception(f"Error parsing sales data: {str(e)}")
    
    def update_excel_sheet(self, workbook, sheet_name, sales_df, product_code):
        try:
            if sheet_name not in workbook.sheetnames:
                self.log_status(f"Warning: Sheet '{sheet_name}' not found in Excel file")
                return
            
            sheet = workbook[sheet_name]
            updates_made = 0
            
            sales_dict = sales_df[product_code].to_dict()
            
            for row in range(2, sheet.max_row + 1):
                outlet = sheet['D' + str(row)].value
                if outlet:
                    outlet_code = str(outlet).split('-')[0] if '-' in str(outlet) else str(outlet)
                    
                    for sales_outlet, sales_value in sales_dict.items():
                        if str(sales_outlet).startswith(outlet_code):
                            if sales_value > 0:
                                sheet['F' + str(row)].value = float(sales_value)
                                updates_made += 1
                                break
            
            self.log_status(f"Updated sheet: {sheet_name} ({updates_made} rows updated)")
            
        except Exception as e:
            raise Exception(f"Error updating sheet {sheet_name}: {str(e)}")
    
    def update_excel(self):
        if not self.excel_path:
            messagebox.showerror("Error", "Please select an Excel file first!")
            return
        
        try:
            sales_df = self.parse_sales_data()
            
            workbook = openpyxl.load_workbook(self.excel_path)
            
            selected_product = self.product_var.get()
            
            if selected_product == 'All':
                for code, sheet_name in self.product_mapping.items():
                    if code in sales_df.columns:
                        self.update_excel_sheet(workbook, sheet_name, sales_df, code)
            else:
                sheet_name = self.product_mapping[selected_product]
                self.update_excel_sheet(workbook, sheet_name, sales_df, selected_product)
            
            workbook.save(self.excel_path)
            messagebox.showinfo("Success", "Excel file updated successfully!")
            self.log_status("Excel file updated successfully")
            
        except Exception as e:
            self.handle_error("Error updating Excel", e)

    def save_analysis(self, fig, sheet_name):
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG files", "*.png"), ("All files", "*.*")],
                initialfile=f"{sheet_name}_analysis.png"
            )
            if file_path:
                fig.savefig(file_path, dpi=300, bbox_inches='tight')
                messagebox.showinfo("Success", "Analysis saved successfully!")
        except Exception as e:
            self.handle_error("Error saving analysis", e)

    def perform_analysis(self):
        try:
            if not self.excel_path:
                messagebox.showerror("Error", "Please select an Excel file first!")
                return
            
            for widget in self.analysis_display.winfo_children():
                widget.destroy()
            
            analysis_notebook = ttk.Notebook(self.analysis_display)
            analysis_notebook.pack(fill="both", expand=True)
            
            selected_product = self.analysis_product_var.get()
            
            if selected_product == "All":
                products_to_analyze = self.product_mapping.items()
            else:
                products_to_analyze = [(selected_product, self.product_mapping[selected_product])]
            
            for code, sheet_name in products_to_analyze:
                product_tab = ttk.Frame(analysis_notebook)
                analysis_notebook.add(product_tab, text=sheet_name)
                
                main_canvas = tk.Canvas(product_tab)
                scrollbar = ttk.Scrollbar(product_tab, orient="vertical", command=main_canvas.yview)
                scrollable_frame = ttk.Frame(main_canvas)
                
                def _on_mousewheel(event):
                    main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                main_canvas.bind_all("<MouseWheel>", _on_mousewheel)
                
                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
                )
                
                window_width = self.root.winfo_width()
                canvas_width = window_width - 50
                main_canvas.configure(width=canvas_width)
                
                main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=canvas_width-20)
                
                wb = openpyxl.load_workbook(self.excel_path, data_only=True)
                sheet = wb[sheet_name]
                
                data = []
                for row in range(2, sheet.max_row + 1):
                    outlet = sheet['D' + str(row)].value
                    target = sheet['E' + str(row)].value
                    sales = sheet['F' + str(row)].value
                    
                    if outlet and target and sales:
                        try:
                            target_val = float(target)
                            sales_val = float(sales)
                            data.append({
                                'Outlet': outlet,
                                'Target': target_val,
                                'Sales': sales_val
                            })
                        except ValueError:
                            continue
                
                if not data:
                    ttk.Label(scrollable_frame, text="No data available for analysis").pack(pady=20)
                    wb.close()
                    continue
                
                df = pd.DataFrame(data)
                
                total_sales = df['Sales'].sum()
                total_target = df['Target'].sum()
                achievement_rate = (total_sales / total_target * 100) if total_target > 0 else 0
                over_target = len(df[df['Sales'] >= df['Target']])
                under_target = len(df[df['Sales'] < df['Target']])
                
                df['Achievement'] = (df['Sales'] / df['Target'] * 100).round(2)
                achievement_rates = df['Achievement']
                
                summary_frame = ttk.LabelFrame(scrollable_frame, text="Performance Summary", padding="10")
                summary_frame.pack(fill="x", padx=5, pady=5)
                
                summary_items = [
                    ("Total Sales:", f"{total_sales:,.2f}"),
                    ("Total Target:", f"{total_target:,.2f}"),
                    ("Overall Achievement Rate:", f"{achievement_rate:.1f}%"),
                    ("Outlets Over Target:", f"{over_target}"),
                    ("Outlets Under Target:", f"{under_target}"),
                    ("Average Sales:", f"{df['Sales'].mean():.2f}"),
                    ("Average Achievement:", f"{df['Achievement'].mean():.1f}%")
                ]
                
                for i, (label, value) in enumerate(summary_items):
                    ttk.Label(summary_frame, text=label, font=("Arial", 10, "bold")).grid(row=i, column=0, sticky="w", padx=5, pady=2)
                    ttk.Label(summary_frame, text=value).grid(row=i, column=1, sticky="w", padx=5, pady=2)
                
                performers_frame = ttk.LabelFrame(scrollable_frame, text="Top/Bottom Performers", padding="10")
                performers_frame.pack(fill="x", padx=5, pady=5)
                
                perf_columns = ttk.Frame(performers_frame)
                perf_columns.pack(fill="x", expand=True)
                
                top_frame = ttk.Frame(perf_columns)
                top_frame.pack(side="left", fill="both", expand=True, padx=5)
                ttk.Label(top_frame, text="Top 5 Performers:", font=("Arial", 10, "bold")).pack(anchor="w")
                for _, row in df.nlargest(5, 'Achievement').iterrows():
                    ttk.Label(top_frame, 
                            text=f"{row['Outlet']}: {row['Achievement']}% Achievement").pack(anchor="w")
                
                bottom_frame = ttk.Frame(perf_columns)
                bottom_frame.pack(side="left", fill="both", expand=True, padx=5)
                ttk.Label(bottom_frame, text="Bottom 5 Performers:", font=("Arial", 10, "bold")).pack(anchor="w")
                for _, row in df.nsmallest(5, 'Achievement').iterrows():
                    ttk.Label(bottom_frame, 
                            text=f"{row['Outlet']}: {row['Achievement']}% Achievement").pack(anchor="w")
                
                fig = plt.figure(figsize=(12, 6))
                
                ax1 = plt.subplot(121)
                x = np.arange(2)
                width = 0.35
                
                bars1 = ax1.bar(x[0], total_target, width, label='Target', color='#2ecc71')
                bars2 = ax1.bar(x[1], total_sales, width, label='Sales', color='#3498db')
                
                ax1.set_ylabel('Amount', fontsize=10)
                ax1.set_title('Total Target vs Sales', pad=20, fontsize=12, fontweight='bold')
                ax1.set_xticks(x)
                ax1.set_xticklabels(['Target', 'Sales'], fontsize=10)
                ax1.legend(fontsize=10)
                
                # Get the maximum value to set proper y-axis limits
                max_value = max(total_target, total_sales)
                ax1.set_ylim(0, max_value * 1.2)  # Add 20% padding at the top
                
                # Position text with offset
                for bars in [bars1, bars2]:
                    for bar in bars:
                        height = bar.get_height()
                        y_pos = height + (max_value * 0.02)  # Position text 2% above bar
                        ax1.text(bar.get_x() + bar.get_width()/2, y_pos,
                                f'{int(height):,}',
                                ha='center', va='bottom',
                                fontsize=10)
                
                ax2 = plt.subplot(122)
                ax2.hist(achievement_rates, bins=20, color='#9b59b6', alpha=0.7, edgecolor='black')
                ax2.set_title('Achievement Rate Distribution', pad=20, fontsize=12, fontweight='bold')
                ax2.set_xlabel('Achievement Rate (%)', fontsize=10)
                ax2.set_ylabel('Number of Outlets', fontsize=10)
                ax2.tick_params(axis='both', which='major', labelsize=9)
                
                plt.tight_layout(pad=3.0)
                
                vis_frame = ttk.LabelFrame(scrollable_frame, text="Visualizations", padding="10")
                vis_frame.pack(fill="x", padx=5, pady=5)
                
                canvas = FigureCanvasTkAgg(fig, vis_frame)
                canvas.draw()
                canvas_widget = canvas.get_tk_widget()
                canvas_widget.configure(height=400)
                canvas_widget.pack(fill="x", padx=5, pady=5)
                
                save_button = ttk.Button(vis_frame, text="Save Analysis", 
                                       command=lambda f=fig, s=sheet_name: self.save_analysis(f, s))
                save_button.pack(pady=5)
                
                scrollbar.pack(side="right", fill="y")
                main_canvas.pack(side="left", fill="both", expand=True)
                
                wb.close()
                plt.close(fig)
            
        except Exception as e:
            self.handle_error("Error performing analysis", e)
            traceback.print_exc()

def main():
    try:
        root = tk.Tk()
        app = SalesUpdateTool(root)
        root.mainloop()
    except Exception as e:
        print(f"Critical error: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    main()